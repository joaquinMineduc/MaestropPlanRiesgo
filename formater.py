import openpyxl
import xlsxwriter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from functions import getYear, coordinate_cell_to_string


wb = openpyxl.load_workbook('Formato_padre_plan.xlsx')

new_workbook = xlsxwriter.Workbook('Formato_padre_plan.xlsx')
new_sheet = new_workbook.add_worksheet()

# Crear formatos
bold_format = new_workbook.add_format({'bold': True})
normal_format = new_workbook.add_format({'bold': False})

years = getYear()

formula1 = f"DATE({years[0]}, {1}, {1})"
formula2 = f"DATE({years[1]}, {12}, {31})"

terms = ['Cumplido:', 'No cumplido:', 'Parcialmente cumplido:']

formula3 = 'IF(LEFT($J6,LEN("Cumplido:"))="Cumplido:", "No aplica", "")'

values = ['Reducir', 'Compartir', 'Aceptar', 'Evitar']
values_str = ','.join(values)

values2 = ['Alta','Media','Baja']
values2_str = ','.join(values2)


formulas = [f'AND(LEFT($J6,{len(term)})="{term}", LEN(SUBSTITUTE($J6," ",""))>{50})' for term in terms]
formula = ' + '.join(formulas)

evidence_formula = '=IF(LEN($K6)<5, NOT(ISBLANK($K6)), TRUE)'


evidence_validation = DataValidation(type = "custom", formula1 = evidence_formula , showErrorMessage = True)

evidence_validation.error = """Estimado, CR. Este campo es obligatorio"""


corrective_formula = '=SI($E5="Media"; "No aplica"; "")'
corrective_validation = DataValidation(type = "custom", formula1 = corrective_formula)



val_goals = DataValidation(type = "custom",
                          formula1 = formula,
                          showErrorMessage = True)

val_goals.error = """El valor debe comenzar con 'Cumplido:', 
'No cumplido:' o 'Parcialmente cumplido:' y tener contenido adicional."""


val_not_null = DataValidation(type = "custom", formula1 = formula3, showErrorMessage = True)
val_not_null.error = "La celda no puede quedar vacía"
val_not_null.errorTitle = "Estimado CR, Hay un error"

# Se crean las validaciones para fechas:
validation_date =  DataValidation(type = "date", 
                                  operator = "between", 
                                  formula1 = formula1, 
                                  formula2 = formula2, 
                                  showErrorMessage = True)

validation_date.error = 'Estimado CR, El valor debe ser una fecha en el formato DD/MM/YYYY'
validation_date.errorTitle = 'El valor debe ser una fecha en el formato DD/MM/YYYY'

validation_estrategy = DataValidation(type = "list", 
                                      formula1 = f'"{values_str}"', 
                                      showErrorMessage = True
                                      )


validation_estrategy.error = """Estimado CR, debe seleccionar sólo 1 
opción de la lista de tipos de estatégias"""

validation_estrategy.errorTitle = "La opción ingresada es inválida"


validation_level = DataValidation(type = "list", 
                                      formula1 = f'"{values2_str}"', 
                                      showErrorMessage = True
                                      )

validation_level.error = """Estimado CR, debe seleccionar sólo 1 
opción de la lista de tipos de estatégias"""

validation_level.errorTitle = "La opción ingresada es inválida"

sheet_names = wb.sheetnames

for sheet in sheet_names:
    
    if sheet == "FINANCIEROS" or sheet == "ESTRATÉGICOS" or sheet == "INSTITUCIONAL":
        print(wb[sheet])
        wb[sheet].add_data_validation(validation_date)
        wb[sheet].add_data_validation(validation_estrategy)
        wb[sheet].add_data_validation(val_goals)
        wb[sheet].add_data_validation(val_not_null)
        wb[sheet].add_data_validation(evidence_validation)
        
        for dim in wb[sheet].iter_rows(min_col = 7, max_col = 7, min_row = 6 , max_row = 50):
            for cell in dim:
                validation_estrategy.add(cell)
        for dim in wb[sheet].iter_rows(min_col = 9, max_col = 9, min_row = 6 , max_row = 50):
            for cell in dim:
               if cell.is_date and cell.value is not None:
                    validation_date.add(cell)
        for dim in wb[sheet].iter_rows(min_col = 10, max_col = 10, min_row = 6 , max_row = 50):
            for cell in dim:
                val_goals.add(cell)
        for dim in wb[sheet].iter_rows(min_col = 12, max_col = 12, min_row = 6 , max_row = 50):
                    for cell in dim:
                        cell.value = f'=IF(LEFT($J{cell.row}, 9)="Cumplido:", "No aplica", "")'
                        val_not_null.add(cell)
        for dim in wb[sheet].iter_rows(min_col = 11, max_col = 11, min_row = 6 , max_row = 50):
                    for cell in dim:
                       evidence_validation.add(cell)
                                 
    if sheet == "SEÑALES DE ALERTA ":
        print(wb[sheet])
        for dim3 in wb[sheet].iter_rows(min_col = 9, max_col = 9, min_row = 7 , max_row = 50):
            for cell in dim3:
                if cell.col_idx == 9 and wb[sheet].title == sheet:
                    wb[sheet].add_data_validation(validation_level)
                    validation_level.add(cell)
                
                    
    if sheet == "S.A NO ASOCIADAS MR":
        print(wb[sheet])
        for dim in wb[sheet].iter_rows(min_col = 5, max_col = 5, min_row = 4 , max_row = 50):
            for cell in dim:
                if cell.col_idx == 5 and wb[sheet].title == sheet:
                    wb[sheet].add_data_validation(validation_level)
                    validation_level.add(cell)
                    
                
                     
        for dim in wb[sheet].iter_rows(min_col = 6, max_col = 6, min_row = 5 , max_row = 50):
            for cell in dim:
                cell.value = f'=IF(OR(LEFT($E{cell.row}, 4)="Baja", LEFT($E{cell.row}, 5)="Media"), "", "No aplica")'
                val_not_null.add(cell)
                                
                      
wb.save("Formato_padre_plan.xlsx")